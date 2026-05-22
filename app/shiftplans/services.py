"""Shift planning generation and validation services."""

import json
import logging
from datetime import date, datetime, timedelta
from io import BytesIO
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

from app.domain_models.common import utc_now
from app.extensions import db
from app.models import (
    Employee,
    EmployeeMachineQualification,
    Machine,
    ShiftPlan,
    ShiftPlanCoverageSlot,
    ShiftPlanEntry,
    VacationRequest,
)
from app.permissions import has_employee_access
from app.shiftplans.generator import build_local_shift_entries, build_local_shift_plan
from app.shiftplans.templates import (
    SHIFT_TEMPLATE_ALIASES,
    SHIFT_TEMPLATES,
    get_shift_model_template,
    normalize_template_value,
    resolve_shift_template,
)

SHIFT_WINDOWS = {
    shift.key: (shift.start_time, shift.end_time)
    for shift in get_shift_model_template("three_shift").shifts
}

SHIFT_LABELS = {
    "frueh": "Frueh",
    "früh": "Frueh",
    "spaet": "Spaet",
    "spät": "Spaet",
    "nacht": "Nacht",
    "frei": "Frei",
    "urlaub": "Urlaub",
}
logger = logging.getLogger(__name__)


def parse_date(value):
    """Parse an ISO date string or default to today."""
    if not value:
        return date.today()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError("start_date must use YYYY-MM-DD") from exc


def parse_days(value):
    """Parse and clamp the shift plan duration in days."""
    try:
        days = int(value or 7)
    except (TypeError, ValueError) as exc:
        raise ValueError("days must be a number") from exc
    return min(max(days, 1), 31)


def department_employees(department):
    """Return employees assigned to a given department for shift planning."""
    return (
        Employee.query.filter(Employee.department.ilike(f"%{department}%"))
        .order_by(Employee.team.asc(), Employee.name.asc())
        .all()
    )


def selected_machines_for_request(data):
    """Return requested machines or all machines for legacy payloads."""
    all_machines = Machine.query.order_by(Machine.name.asc()).all()
    if "machine_ids" not in data:
        return all_machines, None
    raw_machine_ids = data.get("machine_ids")
    if not isinstance(raw_machine_ids, list) or not raw_machine_ids:
        return None, {"error": "Bitte mindestens eine Maschine auswaehlen."}
    try:
        requested_ids = {int(machine_id) for machine_id in raw_machine_ids}
    except (TypeError, ValueError):
        return None, {"error": "machine_ids muessen gueltige Maschinen-IDs enthalten."}
    machine_by_id = {machine.id: machine for machine in all_machines}
    missing_ids = sorted(requested_ids - set(machine_by_id))
    if missing_ids:
        return None, {
            "error": (
                "Unbekannte Maschine(n): "
                + ", ".join(str(machine_id) for machine_id in missing_ids)
            )
        }
    return [machine_by_id[machine_id] for machine_id in sorted(requested_ids)], None


def production_employees():
    """Backward-compatible wrapper: return production employees."""
    return department_employees("Produktion")


def employee_payload(employees):
    """Build the compact employee payload sent to the planner."""
    return [
        {
            "id": employee.id,
            "name": employee.name,
            "team": employee.team,
            "shift_model": employee.shift_model,
            "current_shift": employee.current_shift,
            "qualifications": employee.qualifications,
            "favorite_machine": employee.favorite_machine,
        }
        for employee in employees
    ]


def parse_time(value):
    """Parse a HH:MM time value."""
    return datetime.strptime(value, "%H:%M").time()


def hours_between(start, end):
    """Calculate shift length in hours, supporting overnight shifts."""
    start_time = parse_time(start)
    end_time = parse_time(end)
    start_dt = datetime.combine(date.today(), start_time)
    end_dt = datetime.combine(date.today(), end_time)
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)
    return (end_dt - start_dt).total_seconds() / 3600


def shift_datetimes(work_date, start, end):
    """Return start and end datetimes for one shift entry."""
    start_dt = datetime.combine(work_date, parse_time(start))
    end_dt = datetime.combine(work_date, parse_time(end))
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)
    return start_dt, end_dt


def local_shift_entries(
    start_date,
    days,
    rhythm,
    employees,
    machines,
    unavailable=None,
    shift_model_value=None,
    preferences="",
):
    """Build a fair deterministic fallback plan without calling OpenAI.

    Uses a minimum-shift-count selection instead of round-robin so that
    no employee accumulates significantly more shifts than others.
    """
    return build_local_shift_entries(
        start_date,
        days,
        shift_model_value or rhythm,
        employees,
        machines,
        unavailable=unavailable,
        respect_active_weekdays=bool(shift_model_value),
        preferences=preferences,
    )


def local_shift_plan(
    start_date,
    days,
    rhythm,
    employees,
    machines,
    unavailable=None,
    shift_model_value=None,
    preferences="",
):
    """Build local entries with warnings and visible undercoverage slots."""
    return build_local_shift_plan(
        start_date,
        days,
        shift_model_value or rhythm,
        employees,
        machines,
        unavailable=unavailable,
        respect_active_weekdays=bool(shift_model_value),
        preferences=preferences,
    )


def _pick_fairest_employee(employees, shift_count, work_date, unavailable, assigned_today):
    """Return the available employee with the fewest shifts assigned so far."""
    blocked = unavailable.get(work_date, set())
    candidates = [
        emp for emp in employees if emp.id not in blocked and emp.id not in assigned_today
    ]
    if not candidates:
        return None
    return min(candidates, key=lambda emp: (shift_count[emp.id], emp.id))


def parse_vacation_entries(data, employees, start_date, days):
    """Parse vacation payloads into shift plan entries and unavailable dates."""
    employee_ids = {employee.id for employee in employees}
    vacation_entries = []
    unavailable = {}
    raw_vacations = data.get("vacations") or []
    if not isinstance(raw_vacations, list):
        raise ValueError("vacations must be a list")

    for raw_vacation in raw_vacations:
        if not isinstance(raw_vacation, dict):
            raise ValueError("vacations entries must be objects")
        try:
            employee_id = int(raw_vacation["employee_id"])
            work_date = parse_date(raw_vacation["date"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError("vacations require employee_id and date") from exc
        if employee_id not in employee_ids:
            raise ValueError("vacations contain an unknown production employee")
        if work_date < start_date or work_date >= start_date + timedelta(days=days):
            raise ValueError("vacation date must be within the shift plan range")
        unavailable.setdefault(work_date, set()).add(employee_id)
        vacation_entries.append(
            {
                "employee_id": employee_id,
                "machine_id": None,
                "work_date": work_date,
                "shift": "Urlaub",
                "start_time": "",
                "end_time": "",
                "notes": str(raw_vacation.get("notes") or "Urlaub")[:500],
            }
        )
    return vacation_entries, unavailable


def remove_unavailable_work_entries(entries, unavailable):
    """Return work entries excluding employees blocked by vacation."""
    filtered_entries = []
    for entry in entries:
        try:
            employee_id = int(entry["employee_id"])
            work_date = parse_date(entry["work_date"])
        except (KeyError, TypeError, ValueError):
            filtered_entries.append(entry)
            continue
        if employee_id in unavailable.get(work_date, set()):
            continue
        filtered_entries.append(entry)
    return filtered_entries


def validate_arbzg(entries):
    """Check Arbeitszeitgesetz (ArbZG) rules on a list of entry dicts.

    Returns a list of warning dicts for soft violations.
    Raises ValueError only for hard violations that make the plan invalid.

    Hard errors (plan rejected):
    - One shift per day per employee
    - Max 10 h per shift (§3 Satz 2 — absolute limit)
    - Min 11 h rest between consecutive shifts (§5)

    Soft warnings (plan allowed, admin sees hint):
    - >48 h/week (§3 — average over reference period, not a per-week hard cap)
    - >6 consecutive working days (§9 — depends on employer agreement)
    """
    work_shifts = {"Frueh", "Spaet", "Nacht"}
    warnings = []

    # Collect work entries per employee
    by_employee = {}
    seen_day = set()
    for entry in entries:
        shift = str(entry.get("shift") or "")
        if shift not in work_shifts:
            continue
        emp_id = int(entry["employee_id"])
        work_date = entry["work_date"]
        if isinstance(work_date, str):
            work_date = date.fromisoformat(work_date)

        # HARD: One shift per day per employee
        key = (emp_id, work_date)
        if key in seen_day:
            raise ValueError(
                f"Mitarbeiter {emp_id} hat am {work_date.isoformat()} "
                "mehr als eine Schicht geplant (ArbZG §3)."
            )
        seen_day.add(key)
        by_employee.setdefault(emp_id, []).append(
            {
                "work_date": work_date,
                "start_time": entry["start_time"],
                "end_time": entry["end_time"],
            }
        )

    for emp_id, emp_entries in by_employee.items():
        emp_entries_sorted = sorted(emp_entries, key=lambda e: e["work_date"])

        prev_end_dt = None
        prev_date = None
        consecutive = 0
        weekly_hours = {}

        for e in emp_entries_sorted:
            work_date = e["work_date"]
            start_dt, end_dt = shift_datetimes(work_date, e["start_time"], e["end_time"])
            duration = (end_dt - start_dt).total_seconds() / 3600

            # HARD: Max 10 h per shift (ArbZG §3 Satz 2)
            if duration > 10:
                raise ValueError(
                    f"Mitarbeiter {emp_id}: Schicht am {work_date.isoformat()} "
                    f"dauert {duration:.1f}h — max. 10 Stunden erlaubt (ArbZG §3)."
                )

            # HARD: 11 h rest between shifts (ArbZG §5)
            if prev_end_dt is not None:
                rest_hours = (start_dt - prev_end_dt).total_seconds() / 3600
                if rest_hours < 11:
                    raise ValueError(
                        f"Mitarbeiter {emp_id}: Nur {rest_hours:.1f}h Ruhezeit zwischen "
                        f"{prev_end_dt.date().isoformat()} und {work_date.isoformat()} "
                        "— min. 11 Stunden erforderlich (ArbZG §5)."
                    )

            # SOFT: 48 h/week (ArbZG §3 — Durchschnittswert, kein wöchentliches Maximum)
            week_key = work_date.isocalendar()[:2]
            weekly_hours[week_key] = weekly_hours.get(week_key, 0) + duration
            if weekly_hours[week_key] > 48:
                warnings.append(
                    {
                        "type": "arbzg_weekly_hours",
                        "severity": "warning",
                        "message": (
                            f"Mitarbeiter {emp_id}: {weekly_hours[week_key]:.0f}h in KW "
                            f"{week_key[1]} — Richtwert 48h/Woche ueberschritten (ArbZG §3)."
                        ),
                    }
                )

            # SOFT: 6 consecutive working days (ArbZG §9 — Ausnahmen per Tarifvertrag moeglich)
            if prev_date is not None and (work_date - prev_date).days == 1:
                consecutive += 1
            else:
                consecutive = 1
            if consecutive > 6:
                warnings.append(
                    {
                        "type": "arbzg_consecutive_days",
                        "severity": "warning",
                        "message": (
                            f"Mitarbeiter {emp_id}: {consecutive} aufeinanderfolgende "
                            f"Arbeitstage bis {work_date.isoformat()} (Empfehlung: max. 6)."
                        ),
                    }
                )

            prev_end_dt = end_dt
            prev_date = work_date

    return warnings


def validate_entries(entries, employees, machines, start_date, days):
    """Validate generated entries before they are persisted."""
    employee_ids = {employee.id for employee in employees}
    machine_ids = {machine.id for machine in machines}
    validated = []

    for entry in entries:
        try:
            employee_id = int(entry["employee_id"])
            machine_id = int(entry["machine_id"]) if entry.get("machine_id") else None
            work_date = parse_date(entry["work_date"])
            shift = str(entry.get("shift") or "").strip()[:80]
            start_time = str(entry["start_time"])[:5]
            end_time = str(entry["end_time"])[:5]
        except (KeyError, TypeError, ValueError):
            continue

        if employee_id not in employee_ids:
            continue
        if machine_id and machine_id not in machine_ids:
            continue
        if work_date < start_date or work_date >= start_date + timedelta(days=days):
            continue
        if normalize_shift_name(shift) == "Urlaub":
            validated.append(
                {
                    "employee_id": employee_id,
                    "machine_id": None,
                    "work_date": work_date,
                    "shift": "Urlaub",
                    "start_time": "",
                    "end_time": "",
                    "notes": str(entry.get("notes") or "Urlaub")[:500],
                }
            )
            continue
        if hours_between(start_time, end_time) > 8:
            continue

        validated.append(
            {
                "employee_id": employee_id,
                "machine_id": machine_id,
                "work_date": work_date,
                "shift": normalize_shift_name(shift) or "Schicht",
                "start_time": start_time,
                "end_time": end_time,
                "notes": str(entry.get("notes") or "")[:500],
            }
        )

    return validated


def analyze_shift_plan(entries, employees, machines):
    """Return conflicts and coverage information for generated shift entries."""
    coverage_summary = empty_coverage_summary()
    conflicts = detect_shift_plan_conflicts(
        entries,
        employees,
        machines,
        coverage_summary=coverage_summary,
    )
    return conflicts[:50], coverage_summary


def detect_shift_plan_conflicts(
    entries,
    employees,
    machines,
    coverage_summary=None,
    include_coverage=True,
):
    """Return structured conflicts for shift plan entries."""
    normalized_entries = normalize_conflict_entries(entries)
    employee_by_id = {employee.id: employee for employee in employees}
    machine_by_id = {machine.id: machine for machine in machines}
    conflicts = []
    conflicts.extend(detect_duplicate_assignments(normalized_entries, employee_by_id))
    conflicts.extend(detect_vacation_conflicts(normalized_entries, employee_by_id))
    conflicts.extend(
        detect_missing_machine_qualifications(
            normalized_entries,
            employee_by_id,
            machine_by_id,
        )
    )
    conflicts.extend(detect_rest_time_conflicts(normalized_entries, employee_by_id))
    conflicts.extend(detect_weekly_hours_conflicts(normalized_entries, employee_by_id))
    conflicts.extend(detect_consecutive_day_conflicts(normalized_entries, employee_by_id))
    if include_coverage:
        conflicts.extend(
            update_coverage_summary(
                normalized_entries,
                machines,
                coverage_summary or empty_coverage_summary(),
            )
        )
    return conflicts


def empty_coverage_summary():
    """Return an empty coverage summary payload."""
    return {
        "required_slots": 0,
        "assigned_slots": 0,
        "undercovered": 0,
        "machines": {},
    }


def normalize_conflict_entries(entries):
    """Return dict-based entries for conflict detection."""
    normalized = []
    for entry in entries:
        if isinstance(entry, ShiftPlanEntry):
            normalized.append(
                {
                    "id": entry.id,
                    "employee_id": entry.employee_id,
                    "machine_id": entry.machine_id,
                    "work_date": entry.work_date,
                    "shift": entry.shift,
                    "start_time": entry.start_time,
                    "end_time": entry.end_time,
                    "notes": entry.notes,
                }
            )
            continue
        item = dict(entry)
        if isinstance(item.get("work_date"), str):
            try:
                item["work_date"] = date.fromisoformat(item["work_date"])
            except ValueError:
                pass
        normalized.append(item)
    return normalized


def detect_duplicate_assignments(entries, employee_by_id):
    """Return warnings for employees assigned more than once in a shift window."""
    seen = {}
    warnings = []
    for entry in entries:
        if not entry.get("start_time") or not entry.get("end_time"):
            continue
        key = (
            entry["employee_id"],
            entry["work_date"],
        )
        seen.setdefault(key, []).append(entry)
    for key, key_entries in seen.items():
        if len(key_entries) <= 1:
            continue
        employee = employee_by_id.get(key[0])
        warnings.append(
            {
                "type": "duplicate_assignment",
                "severity": "critical",
                "employee_id": key[0],
                "work_date": key[1].isoformat(),
                "entry_ids": [entry.get("id") for entry in key_entries if entry.get("id")],
                "message": (
                    f"{employee.name if employee else 'Mitarbeiter'} ist am "
                    f"{key[1].isoformat()} mehrfach geplant."
                ),
            }
        )
    return warnings


def detect_vacation_conflicts(entries, employee_by_id):
    """Return conflicts for work entries during approved vacations."""
    employee_ids = {entry.get("employee_id") for entry in entries if entry.get("employee_id")}
    dates = [
        entry.get("work_date") for entry in entries if isinstance(entry.get("work_date"), date)
    ]
    if not employee_ids or not dates:
        return []
    approved_vacations = VacationRequest.query.filter(
        VacationRequest.employee_id.in_(employee_ids),
        VacationRequest.status == "approved",
        VacationRequest.start_date <= max(dates),
        VacationRequest.end_date >= min(dates),
    ).all()
    vacation_days = {}
    for vacation in approved_vacations:
        current = vacation.start_date
        while current <= vacation.end_date:
            vacation_days[(vacation.employee_id, current)] = vacation
            current += timedelta(days=1)

    conflicts = []
    for entry in entries:
        if not is_work_entry(entry):
            continue
        vacation = vacation_days.get((entry.get("employee_id"), entry.get("work_date")))
        if not vacation:
            continue
        employee = employee_by_id.get(entry.get("employee_id"))
        conflicts.append(
            {
                "type": "vacation_conflict",
                "severity": "critical",
                "employee_id": entry.get("employee_id"),
                "entry_id": entry.get("id"),
                "vacation_request_id": vacation.id,
                "work_date": entry["work_date"].isoformat(),
                "message": (
                    f"{employee.name if employee else 'Mitarbeiter'} ist am "
                    f"{entry['work_date'].isoformat()} trotz genehmigtem Urlaub geplant."
                ),
            }
        )
    return conflicts


def detect_rest_time_conflicts(entries, employee_by_id):
    """Return warnings for entries with less than 11 hours rest time."""
    warnings = []
    by_employee = {}
    for entry in entries:
        if not is_work_entry(entry):
            continue
        by_employee.setdefault(entry["employee_id"], []).append(entry)
    for employee_id, employee_entries in by_employee.items():
        sorted_entries = sorted(
            employee_entries,
            key=lambda item: shift_datetimes(
                item["work_date"],
                item["start_time"],
                item["end_time"],
            )[0],
        )
        previous_end = None
        for entry in sorted_entries:
            start_dt, end_dt = shift_datetimes(
                entry["work_date"],
                entry["start_time"],
                entry["end_time"],
            )
            if previous_end:
                rest_hours = (start_dt - previous_end).total_seconds() / 3600
                if rest_hours < 11:
                    employee = employee_by_id.get(employee_id)
                    warnings.append(
                        {
                            "type": "rest_time",
                            "severity": "critical",
                            "employee_id": employee_id,
                            "entry_id": entry.get("id"),
                            "work_date": entry["work_date"].isoformat(),
                            "message": (
                                f"{employee.name if employee else 'Mitarbeiter'} "
                                f"hat nur {round(rest_hours, 1)}h Ruhezeit."
                            ),
                        }
                    )
            previous_end = end_dt
    return warnings


def detect_missing_machine_qualifications(entries, employee_by_id, machine_by_id):
    """Return conflicts when structured machine qualification is missing."""
    conflicts = []
    employee_ids = {entry.get("employee_id") for entry in entries if entry.get("employee_id")}
    machine_ids = {entry.get("machine_id") for entry in entries if entry.get("machine_id")}
    qualifications = EmployeeMachineQualification.query.filter(
        EmployeeMachineQualification.employee_id.in_(employee_ids or {0}),
        EmployeeMachineQualification.machine_id.in_(machine_ids or {0}),
    ).all()
    qualification_map = {
        (qualification.employee_id, qualification.machine_id): qualification
        for qualification in qualifications
    }
    if not qualifications and entries_use_legacy_qualification_mode(entries):
        return []
    for entry in entries:
        machine_id = entry.get("machine_id")
        if not machine_id or not is_work_entry(entry):
            continue
        employee = employee_by_id.get(entry["employee_id"])
        machine = machine_by_id.get(machine_id)
        if not employee or not machine:
            continue
        qualification = qualification_map.get((employee.id, machine_id))
        if qualification and qualification.is_valid_for(entry["work_date"]):
            continue
        conflicts.append(
            {
                "type": "missing_qualification",
                "severity": "critical",
                "employee_id": employee.id,
                "machine_id": machine_id,
                "entry_id": entry.get("id"),
                "work_date": entry["work_date"].isoformat(),
                "message": (
                    f"{employee.name} hat keine gueltige Maschinenfreigabe " f"fuer {machine.name}."
                ),
            }
        )
    return conflicts[:50]


def entries_use_legacy_qualification_mode(entries):
    """Return whether generated entries intentionally used legacy qualification data."""
    work_entries = [entry for entry in entries if is_work_entry(entry)]
    if not work_entries:
        return False
    return all("Legacy-" in str(entry.get("notes") or "") for entry in work_entries)


def detect_weekly_hours_conflicts(entries, employee_by_id):
    """Return conflicts for weekly working hours above 48 hours."""
    weekly_hours = {}
    for entry in entries:
        if not is_work_entry(entry):
            continue
        week_key = (entry["employee_id"], entry["work_date"].isocalendar()[:2])
        weekly_hours[week_key] = weekly_hours.get(week_key, 0) + hours_between(
            entry["start_time"],
            entry["end_time"],
        )
    conflicts = []
    for (employee_id, week_key), hours in weekly_hours.items():
        if hours <= 48:
            continue
        employee = employee_by_id.get(employee_id)
        conflicts.append(
            {
                "type": "weekly_hours",
                "severity": "warning",
                "employee_id": employee_id,
                "calendar_week": week_key[1],
                "hours": round(hours, 2),
                "message": (
                    f"{employee.name if employee else 'Mitarbeiter'} ist in KW "
                    f"{week_key[1]} mit {hours:.1f}h geplant."
                ),
            }
        )
    return conflicts


def detect_consecutive_day_conflicts(entries, employee_by_id):
    """Return conflicts for more than six consecutive working days."""
    by_employee = {}
    for entry in entries:
        if is_work_entry(entry):
            by_employee.setdefault(entry["employee_id"], set()).add(entry["work_date"])
    conflicts = []
    for employee_id, work_dates in by_employee.items():
        consecutive = 0
        previous_date = None
        for work_date in sorted(work_dates):
            if previous_date and (work_date - previous_date).days == 1:
                consecutive += 1
            else:
                consecutive = 1
            if consecutive > 6:
                employee = employee_by_id.get(employee_id)
                conflicts.append(
                    {
                        "type": "consecutive_days",
                        "severity": "warning",
                        "employee_id": employee_id,
                        "work_date": work_date.isoformat(),
                        "days": consecutive,
                        "message": (
                            f"{employee.name if employee else 'Mitarbeiter'} ist "
                            f"{consecutive} Tage in Folge geplant."
                        ),
                    }
                )
            previous_date = work_date
    return conflicts


def detect_vacation_assignment_warnings(entries, vacation_entries, employee_by_id):
    """Return warnings when a vacation day still contains working entries."""
    vacation_days = {(entry["employee_id"], entry["work_date"]) for entry in vacation_entries}
    warnings = []
    for entry in entries:
        if normalize_shift_name(entry.get("shift")) == "Urlaub":
            continue
        key = (entry["employee_id"], entry["work_date"])
        if key not in vacation_days:
            continue
        employee = employee_by_id.get(entry["employee_id"])
        warnings.append(
            {
                "type": "vacation_conflict",
                "severity": "critical",
                "message": (
                    f"{employee.name if employee else 'Mitarbeiter'} ist am "
                    f"{entry['work_date'].isoformat()} trotz Urlaub geplant."
                ),
            }
        )
    return warnings


def update_coverage_summary(entries, machines, coverage_summary):
    """Update coverage summary and return undercoverage warnings."""
    warnings = []
    assigned = {}
    work_dates = sorted(
        {entry["work_date"] for entry in entries if isinstance(entry.get("work_date"), date)}
    )
    work_shifts = sorted(
        {entry["shift"] for entry in entries if is_work_entry(entry)} or {"Frueh", "Spaet"}
    )
    for entry in entries:
        if not entry.get("machine_id") or not is_work_entry(entry):
            continue
        key = (entry["machine_id"], entry["work_date"], entry["shift"])
        assigned[key] = assigned.get(key, 0) + 1

    for machine in machines:
        machine_required = 0
        machine_assigned = 0
        for work_date in work_dates:
            for shift in work_shifts:
                key = (machine.id, work_date, shift)
                count = assigned.get(key, 0)
                machine_required += machine.required_employees
                machine_assigned += count
                if count < machine.required_employees:
                    coverage_summary["undercovered"] += 1
                    warnings.append(
                        {
                            "type": "coverage",
                            "severity": "critical",
                            "machine_id": machine.id,
                            "work_date": key[1].isoformat(),
                            "shift": key[2],
                            "required": machine.required_employees,
                            "assigned": count,
                            "message": (
                                f"{machine.name} ist am {key[1].isoformat()} "
                                f"in {key[2]} unterbesetzt."
                            ),
                        }
                    )
        coverage_summary["machines"][machine.name] = {
            "required_slots": machine_required,
            "assigned_slots": machine_assigned,
        }
        coverage_summary["required_slots"] += machine_required
        coverage_summary["assigned_slots"] += machine_assigned
    return warnings


def build_template_coverage_summary(
    entries,
    machines,
    start_date,
    days,
    shift_model_value,
    respect_active_weekdays=True,
):
    """Return coverage summary and undercoverage slots for a template plan."""
    template = resolve_shift_template(shift_model_value)
    normalized_entries = normalize_conflict_entries(entries)
    assigned = {}
    for entry in normalized_entries:
        if not entry.get("machine_id") or not is_work_entry(entry):
            continue
        key = (entry["machine_id"], entry["work_date"], entry["shift"])
        assigned[key] = assigned.get(key, 0) + 1

    coverage_summary = empty_coverage_summary()
    unassigned_slots = []
    for machine in machines:
        machine_required = 0
        machine_assigned = 0
        for day_offset in range(days):
            work_date = start_date + timedelta(days=day_offset)
            if respect_active_weekdays and not template.is_active_on(work_date):
                continue
            for shift_window in template.shifts:
                key = (machine.id, work_date, shift_window.key)
                required = int(machine.required_employees)
                count = assigned.get(key, 0)
                missing = max(0, required - count)
                machine_required += required
                machine_assigned += count
                if missing:
                    coverage_summary["undercovered"] += 1
                    unassigned_slots.append(
                        undercoverage_slot_payload(
                            machine,
                            work_date,
                            shift_window.key,
                            required,
                            count,
                            missing,
                        )
                    )
        coverage_summary["machines"][machine.name] = {
            "required_slots": machine_required,
            "assigned_slots": machine_assigned,
        }
        coverage_summary["required_slots"] += machine_required
        coverage_summary["assigned_slots"] += machine_assigned
    return coverage_summary, unassigned_slots


def undercoverage_slot_payload(machine, work_date, shift, required, assigned, missing):
    """Return a visible undercoverage payload for one machine shift."""
    return {
        "work_date": work_date.isoformat(),
        "shift": shift,
        "machine_id": machine.id if machine else None,
        "machine_name": machine.name if machine else None,
        "required": required,
        "assigned": assigned,
        "missing": missing,
        "reason": "Keine regelkonforme Besetzung moeglich",
        "suggestion": (
            "Maschinenqualifikationen pflegen oder zusaetzliche "
            "Mitarbeitende freigeben"
        ),
    }


def coverage_warnings_from_slots(unassigned_slots):
    """Return critical coverage warnings for visible undercoverage slots."""
    warnings = []
    for slot in unassigned_slots:
        machine_text = f" an {slot['machine_name']}" if slot.get("machine_name") else ""
        warnings.append(
            {
                "type": "coverage",
                "severity": "critical",
                "machine_id": slot.get("machine_id"),
                "machine_name": slot.get("machine_name"),
                "work_date": slot.get("work_date"),
                "shift": slot.get("shift"),
                "required": slot.get("required"),
                "assigned": slot.get("assigned"),
                "missing": slot.get("missing"),
                "message": (
                    f"Keine regelkonforme Besetzung am {slot.get('work_date')} "
                    f"fuer {slot.get('shift')}{machine_text}."
                ),
            }
        )
    return warnings


def is_work_entry(entry):
    """Return whether an entry represents planned work."""
    shift = normalize_shift_name(entry.get("shift"))
    return bool(
        shift not in {"Urlaub", "Frei", ""} and entry.get("start_time") and entry.get("end_time")
    )


def generated_work_entries(entries):
    """Return persisted entry payloads that represent generated work."""
    return [entry for entry in entries if is_work_entry(entry)]


def normalize_shift_name(value):
    """Return a supported display shift name for common German spellings."""
    normalized = str(value or "").strip().lower()
    return SHIFT_LABELS.get(normalized, str(value or "").strip())


def normalize_preferences(value):
    """Return preferences as bounded text for storage and scoring."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict | list):
        return json.dumps(value, ensure_ascii=True, sort_keys=True)
    return str(value).strip()


def is_known_shift_model_value(value):
    """Return whether a value is a supported model key or explicit alias."""
    raw_value = str(value or "").strip()
    if raw_value in SHIFT_TEMPLATES:
        return True
    normalized = normalize_template_value(value)
    normalized_aliases = {
        normalize_template_value(alias): key
        for alias, key in SHIFT_TEMPLATE_ALIASES.items()
    }
    return normalized in SHIFT_TEMPLATES or normalized in normalized_aliases


def resolve_explicit_shift_model_key(data):
    """Return the canonical explicit shift model key or a validation error."""
    if "shift_model_key" in data:
        explicit_value = data.get("shift_model_key")
    elif "shift_model" in data:
        explicit_value = data.get("shift_model")
    else:
        return None, None
    if not str(explicit_value or "").strip():
        return None, {
            "error": (
                "shift_model_key darf nicht leer sein. Bitte ein Modell aus "
                "/api/v1/shiftplans/models verwenden."
            )
        }
    if not is_known_shift_model_value(explicit_value):
        return None, {
            "error": (
                "Unbekanntes Schichtmodell. Bitte ein Modell aus "
                "/api/v1/shiftplans/models verwenden."
            )
        }
    raw_value = str(explicit_value or "").strip()
    if raw_value in SHIFT_TEMPLATES:
        return raw_value, None
    normalized = normalize_template_value(explicit_value)
    normalized_aliases = {
        normalize_template_value(alias): key
        for alias, key in SHIFT_TEMPLATE_ALIASES.items()
    }
    canonical_key = normalized_aliases.get(normalized, normalized)
    return canonical_key, None


def build_shift_plan_draft(data):
    """Build a validated shift plan draft without persisting database rows."""
    department = (data.get("department") or "").strip()
    if not department:
        return None, {"error": "Abteilung ist erforderlich"}, 400

    try:
        start_date = parse_date(data.get("start_date"))
        days = parse_days(data.get("days"))
    except ValueError as exc:
        return None, {"error": str(exc)}, 400
    shift_model_value, shift_model_error = resolve_explicit_shift_model_key(data)
    if shift_model_error:
        return None, shift_model_error, 400
    rhythm = shift_model_value or data.get("rhythm", "2-Schicht Rhythmus")
    preferences = normalize_preferences(data.get("preferences", ""))
    title = data.get("title") or f"Schichtplan {department} ab {start_date.isoformat()}"

    employees = department_employees(department)
    machines, machine_error = selected_machines_for_request(data)
    if machine_error:
        return None, machine_error, 400
    if not employees:
        return None, {"error": f"Keine Mitarbeitenden in Abteilung '{department}' gefunden"}, 400

    try:
        vacation_entries, unavailable = parse_vacation_entries(
            data,
            employees,
            start_date,
            days,
        )
    except ValueError as exc:
        return None, {"error": str(exc)}, 400
    import_approved_vacations(employees, vacation_entries, unavailable, start_date, days)

    raw_entries, planning_warnings, generator_unassigned = local_shift_plan(
        start_date,
        days,
        rhythm,
        employees,
        machines,
        unavailable,
        shift_model_value=shift_model_value,
        preferences=preferences,
    )
    entries = validate_entries(
        raw_entries + vacation_entries,
        employees,
        machines,
        start_date,
        days,
    )
    if not entries and not planning_warnings:
        return None, {"error": "Es konnte kein gueltiger Schichtplan erzeugt werden"}, 400
    if not generated_work_entries(entries):
        return None, {
            "error": (
                "Kein Plan erzeugt. Bitte Maschinenqualifikationen pflegen "
                "oder Mitarbeiterdaten pruefen."
            ),
            "warnings": planning_warnings,
            "unassigned_slots": generator_unassigned,
        }, 422

    try:
        arbzg_warnings = validate_arbzg(
            [entry for entry in entries if entry.get("shift") not in ("Urlaub", "Frei", "")]
        )
    except ValueError as exc:
        return None, {"error": str(exc)}, 422

    coverage_summary, unassigned_slots = build_template_coverage_summary(
        entries,
        machines,
        start_date,
        days,
        shift_model_value or rhythm,
        respect_active_weekdays=bool(shift_model_value),
    )
    warnings = detect_shift_plan_conflicts(
        entries,
        employees,
        machines,
        coverage_summary=empty_coverage_summary(),
        include_coverage=False,
    )[:50]
    employee_by_id = {employee.id: employee for employee in employees}
    warnings.extend(coverage_warnings_from_slots(unassigned_slots))
    warnings.extend(arbzg_warnings)
    warnings.extend(
        detect_vacation_assignment_warnings(entries, vacation_entries, employee_by_id)
    )
    return {
        "title": title,
        "department": department,
        "start_date": start_date,
        "days": days,
        "rhythm": rhythm,
        "shift_model_value": shift_model_value,
        "preferences": preferences,
        "employees": employees,
        "machines": machines,
        "entries": entries,
        "warnings": warnings,
        "coverage_summary": coverage_summary,
        "unassigned_slots": unassigned_slots,
        "fairness_summary": fairness_summary(entries, employees),
    }, None, 200


def import_approved_vacations(employees, vacation_entries, unavailable, start_date, days):
    """Append approved vacation requests to generated planning inputs."""
    try:
        employee_ids = [employee.id for employee in employees]
        period_end = start_date + timedelta(days=days - 1)
        approved_vacations = VacationRequest.query.filter(
            VacationRequest.employee_id.in_(employee_ids),
            VacationRequest.status == "approved",
            VacationRequest.start_date <= period_end,
            VacationRequest.end_date >= start_date,
        ).all()
        for vacation in approved_vacations:
            current_date = max(vacation.start_date, start_date)
            while current_date <= min(vacation.end_date, period_end):
                if current_date.weekday() < 5:
                    key = (vacation.employee_id, current_date)
                    existing_keys = {
                        (entry["employee_id"], entry["work_date"])
                        for entry in vacation_entries
                        if isinstance(entry["work_date"], type(current_date))
                    }
                    if key not in existing_keys:
                        vacation_entries.append(
                            {
                                "employee_id": vacation.employee_id,
                                "machine_id": None,
                                "work_date": current_date,
                                "shift": "Urlaub",
                                "start_time": "",
                                "end_time": "",
                                "notes": f"Genehmigter Urlaub (Antrag #{vacation.id})",
                            }
                        )
                        unavailable.setdefault(current_date, set()).add(vacation.employee_id)
                current_date += timedelta(days=1)
    except Exception:
        logger.exception("Vacation import failed; continuing plan generation")


def preview_shift_plan(data):
    """Return a generated shift plan preview without saving it."""
    draft, error, status = build_shift_plan_draft(data)
    if error:
        return None, error, status
    return preview_payload(draft), None, 200


def preview_payload(draft):
    """Return API payload for a generated shift plan draft."""
    return {
        "id": None,
        "is_preview": True,
        "title": draft["title"],
        "start_date": draft["start_date"].isoformat(),
        "days": draft["days"],
        "rhythm": draft["rhythm"],
        "preferences": draft["preferences"],
        "notes": "Vorschau ohne Speicherung.",
        "department": draft["department"],
        "status": "preview",
        "coverage_percent": coverage_percent(draft["coverage_summary"]),
        "conflict_count": conflict_summary(draft["warnings"])["total"],
        "critical_conflict_count": conflict_summary(draft["warnings"])["critical"],
        "change_count": 0,
        "created_by": None,
        "entries": serialize_entry_payloads(
            draft["entries"],
            draft["employees"],
            draft["machines"],
        ),
        "warnings": draft["warnings"],
        "conflicts": draft["warnings"],
        "coverage_summary": draft["coverage_summary"],
        "unassigned_slots": draft["unassigned_slots"],
        "fairness_summary": draft["fairness_summary"],
        "required_employee_count": draft["coverage_summary"].get("required_slots", 0),
    }


def coverage_percent(coverage_summary):
    """Return a percentage for assigned versus required slots."""
    required_slots = coverage_summary.get("required_slots") or 0
    assigned_slots = coverage_summary.get("assigned_slots") or 0
    return round((assigned_slots / required_slots) * 100, 2) if required_slots else 0


def serialize_entry_payloads(entries, employees, machines):
    """Return plan-like JSON entries for preview payloads."""
    employee_by_id = {employee.id: employee for employee in employees}
    machine_by_id = {machine.id: machine for machine in machines}
    payload = []
    for index, entry in enumerate(entries, start=1):
        employee = employee_by_id.get(entry["employee_id"])
        machine = machine_by_id.get(entry.get("machine_id"))
        payload.append(
            {
                "id": None,
                "preview_id": index,
                "employee": employee.to_dict("confidential") if employee else None,
                "machine": machine.to_dict() if machine else None,
                "work_date": (
                    entry["work_date"].isoformat()
                    if isinstance(entry["work_date"], date)
                    else entry["work_date"]
                ),
                "shift": entry["shift"],
                "start_time": entry["start_time"],
                "end_time": entry["end_time"],
                "notes": entry.get("notes", ""),
                "created_at": None,
            }
        )
    return payload


def fairness_summary(entries, employees):
    """Return compact fairness counters for a generated draft."""
    by_employee = {
        employee.id: {
            "employee_id": employee.id,
            "name": employee.name,
            "shifts": 0,
            "night_shifts": 0,
            "weekend_shifts": 0,
            "hours": 0.0,
        }
        for employee in employees
    }
    for entry in entries:
        if not is_work_entry(entry):
            continue
        employee_id = int(entry["employee_id"])
        item = by_employee.get(employee_id)
        if not item:
            continue
        work_date = entry["work_date"]
        if isinstance(work_date, str):
            work_date = date.fromisoformat(work_date)
        item["shifts"] += 1
        item["hours"] += hours_between(entry["start_time"], entry["end_time"])
        if entry["shift"] == "Nacht":
            item["night_shifts"] += 1
        if work_date.weekday() >= 5:
            item["weekend_shifts"] += 1
    return {"employees": list(by_employee.values())}


def persist_shift_plan_from_draft(data, user=None):
    """Persist a generated shift plan draft and return the saved plan."""
    draft, error, status = build_shift_plan_draft(data)
    if error:
        return None, error, status

    notes = (
        "Regelbasierter Generator genutzt. Ungueltige Zuweisungen werden "
        "vor dem Speichern blockiert."
    )
    plan = ShiftPlan(
        title=draft["title"],
        start_date=draft["start_date"],
        days=draft["days"],
        rhythm=draft["rhythm"],
        preferences=draft["preferences"],
        notes=notes,
        department=draft["department"],
        created_by=user.id if user else None,
    )
    summary = conflict_summary(draft["warnings"])
    plan.coverage_percent = coverage_percent(draft["coverage_summary"])
    plan.conflict_count = summary["total"]
    plan.critical_conflict_count = summary["critical"]
    db.session.add(plan)
    db.session.flush()

    for entry in draft["entries"]:
        db.session.add(ShiftPlanEntry(plan=plan, **entry))
    db.session.flush()
    replace_plan_coverage_slots(plan, draft["unassigned_slots"])
    update_employee_rotation_state(draft["employees"])
    db.session.commit()

    plan.warnings = draft["warnings"]
    plan.coverage_summary = draft["coverage_summary"]
    plan.fairness_summary = draft["fairness_summary"]
    return plan, None, 201


def replace_plan_coverage_slots(plan, unassigned_slots):
    """Replace persisted undercoverage slots for a plan."""
    ShiftPlanCoverageSlot.query.filter_by(plan_id=plan.id).delete()
    for slot in unassigned_slots:
        db.session.add(
            ShiftPlanCoverageSlot(
                plan_id=plan.id,
                machine_id=slot.get("machine_id"),
                work_date=parse_date(slot.get("work_date")),
                shift=str(slot.get("shift") or "")[:80],
                required=int(slot.get("required") or 0),
                assigned=int(slot.get("assigned") or 0),
                missing=int(slot.get("missing") or 0),
                reason=str(slot.get("reason") or "")[:240],
                suggestion=str(slot.get("suggestion") or "")[:240],
            )
        )


def refresh_plan_coverage_slots(plan):
    """Recalculate and persist visible undercoverage slots for a plan."""
    machines = machines_for_plan(plan)
    coverage_summary, unassigned_slots = build_template_coverage_summary(
        list(plan.entries),
        machines,
        plan.start_date,
        plan.days,
        plan.rhythm,
        respect_active_weekdays=is_known_shift_model_value(plan.rhythm),
    )
    replace_plan_coverage_slots(plan, unassigned_slots)
    return coverage_summary, unassigned_slots


def update_employee_rotation_state(employees):
    """Update last, current, and next planned shift values for employees."""
    today = date.today()
    for employee in employees:
        entries = (
            ShiftPlanEntry.query.filter_by(employee_id=employee.id)
            .filter(ShiftPlanEntry.shift.notin_(("Urlaub", "Frei", "")))
            .order_by(ShiftPlanEntry.work_date.asc(), ShiftPlanEntry.start_time.asc())
            .all()
        )
        past_entries = [entry for entry in entries if entry.work_date < today]
        upcoming_entries = [entry for entry in entries if entry.work_date >= today]
        employee.last_shift = past_entries[-1].shift if past_entries else ""
        employee.current_shift = upcoming_entries[0].shift if upcoming_entries else ""
        employee.next_shift = upcoming_entries[1].shift if len(upcoming_entries) > 1 else ""
        employee.rotation_state_updated_at = utc_now()


def generate_shift_plan(data, user=None):
    """Generate, validate and save a shift plan from request data.

    Args:
        data: Parsed JSON request body.
        user: The User object of the requesting user (for audit trail).

    """
    return persist_shift_plan_from_draft(data, user)

    department = (data.get("department") or "").strip()
    if not department:
        return None, {"error": "Abteilung ist erforderlich"}, 400

    try:
        start_date = parse_date(data.get("start_date"))
        days = parse_days(data.get("days"))
    except ValueError as exc:
        return None, {"error": str(exc)}, 400
    shift_model_value, shift_model_error = resolve_explicit_shift_model_key(data)
    if shift_model_error:
        return None, shift_model_error, 400
    rhythm = shift_model_value or data.get("rhythm", "2-Schicht Rhythmus")
    preferences = normalize_preferences(data.get("preferences", ""))
    title = data.get("title") or f"Schichtplan {department} ab {start_date.isoformat()}"

    employees = department_employees(department)
    machines = Machine.query.order_by(Machine.name.asc()).all()
    if not employees:
        return None, {"error": f"Keine Mitarbeitenden in Abteilung '{department}' gefunden"}, 400
    try:
        vacation_entries, unavailable = parse_vacation_entries(
            data,
            employees,
            start_date,
            days,
        )
    except ValueError as exc:
        return None, {"error": str(exc)}, 400

    # Auto-import approved vacation requests for this period
    try:
        from app.models import VacationRequest

        employee_ids = [e.id for e in employees]
        period_end = start_date + timedelta(days=days - 1)
        approved_vrs = VacationRequest.query.filter(
            VacationRequest.employee_id.in_(employee_ids),
            VacationRequest.status == "approved",
            VacationRequest.start_date <= period_end,
            VacationRequest.end_date >= start_date,
        ).all()
        for vr in approved_vrs:
            d = max(vr.start_date, start_date)
            while d <= min(vr.end_date, period_end):
                if d.weekday() < 5:
                    key = (vr.employee_id, d)
                    if key not in {
                        (e["employee_id"], e["work_date"])
                        for e in vacation_entries
                        if isinstance(e["work_date"], type(d))
                    }:
                        vacation_entries.append(
                            {
                                "employee_id": vr.employee_id,
                                "machine_id": None,
                                "work_date": d,
                                "shift": "Urlaub",
                                "start_time": "",
                                "end_time": "",
                                "notes": f"Genehmigter Urlaub (Antrag #{vr.id})",
                            }
                        )
                        unavailable.setdefault(d, set()).add(vr.employee_id)
                d += timedelta(days=1)
    except Exception:
        logger.exception("Vacation import failed; continuing plan generation")

    raw_entries, planning_warnings = local_shift_entries(
        start_date,
        days,
        rhythm,
        employees,
        machines,
        unavailable,
        shift_model_value=shift_model_value,
        preferences=preferences,
    )
    notes = (
        "Regelbasierter Generator genutzt. Ungueltige Zuweisungen werden "
        "vor dem Speichern blockiert."
    )

    entries = validate_entries(
        raw_entries + vacation_entries,
        employees,
        machines,
        start_date,
        days,
    )
    if not entries and not planning_warnings:
        return None, {"error": "Es konnte kein gueltiger Schichtplan erzeugt werden"}, 400
    if not generated_work_entries(entries):
        return None, {
            "error": (
                "Kein Plan erzeugt. Bitte Maschinenqualifikationen pflegen "
                "oder Mitarbeiterdaten pruefen."
            ),
            "warnings": planning_warnings,
        }, 422

    # Check Arbeitszeitgesetz — hard violations reject plan, soft ones become warnings
    try:
        arbzg_warnings = validate_arbzg(
            [e for e in entries if e.get("shift") not in ("Urlaub", "Frei", "")]
        )
    except ValueError as exc:
        return None, {"error": str(exc)}, 422

    warnings, coverage_summary = analyze_shift_plan(entries, employees, machines)
    employee_by_id = {employee.id: employee for employee in employees}
    warnings.extend(planning_warnings)
    warnings.extend(arbzg_warnings)
    warnings.extend(detect_vacation_assignment_warnings(entries, vacation_entries, employee_by_id))

    plan = ShiftPlan(
        title=title,
        start_date=start_date,
        days=days,
        rhythm=rhythm,
        preferences=preferences,
        notes=notes,
        department=department,
        created_by=user.id if user else None,
    )
    summary = conflict_summary(warnings)
    required_slots = coverage_summary.get("required_slots") or 0
    assigned_slots = coverage_summary.get("assigned_slots") or 0
    plan.coverage_percent = (
        round((assigned_slots / required_slots) * 100, 2) if required_slots else 0
    )
    plan.conflict_count = summary["total"]
    plan.critical_conflict_count = summary["critical"]
    db.session.add(plan)
    db.session.flush()

    for entry in entries:
        db.session.add(ShiftPlanEntry(plan=plan, **entry))

    db.session.commit()
    plan.warnings = warnings
    plan.coverage_summary = coverage_summary
    return plan, None, 201


def conflicts_for_plan(plan):
    """Return structured conflicts and coverage summary for a persisted plan."""
    employees = employees_for_entries(plan.entries)
    machines = machines_for_plan(plan)
    coverage_summary, unassigned_slots = build_template_coverage_summary(
        list(plan.entries),
        machines,
        plan.start_date,
        plan.days,
        plan.rhythm,
        respect_active_weekdays=is_known_shift_model_value(plan.rhythm),
    )
    conflicts = detect_shift_plan_conflicts(
        list(plan.entries),
        employees,
        machines,
        coverage_summary=empty_coverage_summary(),
        include_coverage=False,
    )
    conflicts.extend(coverage_warnings_from_slots(unassigned_slots))
    return {
        "plan_id": plan.id,
        "conflicts": conflicts,
        "summary": conflict_summary(conflicts),
        "coverage_summary": coverage_summary,
        "unassigned_slots": unassigned_slots,
    }


def machines_for_plan(plan):
    """Return machines that belong to a plan's entries or coverage slots."""
    machine_ids = {
        entry.machine_id for entry in plan.entries if entry.machine_id is not None
    }
    machine_ids.update(
        slot.machine_id for slot in plan.coverage_slots if slot.machine_id is not None
    )
    if not machine_ids:
        return Machine.query.order_by(Machine.name.asc()).all()
    return (
        Machine.query.filter(Machine.id.in_(machine_ids))
        .order_by(Machine.name.asc())
        .all()
    )


def validate_shiftplan_payload(data):
    """Validate an ad-hoc shift plan payload or an existing plan id."""
    if data.get("plan_id"):
        plan = db.session.get(ShiftPlan, int(data["plan_id"]))
        if not plan:
            return None, {"error": "Schichtplan nicht gefunden"}, 404
        return conflicts_for_plan(plan), None, 200

    raw_entries = data.get("entries") or []
    if not isinstance(raw_entries, list):
        return None, {"error": "entries must be a list"}, 400

    entries = []
    for raw_entry in raw_entries:
        entry, error = normalize_validation_entry(raw_entry)
        if error:
            return None, {"error": error}, 400
        entries.append(entry)

    employees = employees_for_entries(entries)
    machines = Machine.query.order_by(Machine.name.asc()).all()
    coverage_summary = {
        "required_slots": 0,
        "assigned_slots": 0,
        "undercovered": 0,
        "machines": {},
    }
    conflicts = detect_shift_plan_conflicts(
        entries,
        employees,
        machines,
        coverage_summary=coverage_summary,
    )
    return (
        {
            "plan_id": None,
            "conflicts": conflicts,
            "summary": conflict_summary(conflicts),
            "coverage_summary": coverage_summary,
        },
        None,
        200,
    )


def normalize_validation_entry(raw_entry):
    """Validate one ad-hoc validation entry."""
    if not isinstance(raw_entry, dict):
        return None, "entries must contain objects"
    try:
        return (
            {
                "id": raw_entry.get("id"),
                "employee_id": int(raw_entry["employee_id"]),
                "machine_id": (
                    int(raw_entry["machine_id"]) if raw_entry.get("machine_id") else None
                ),
                "work_date": parse_date(raw_entry["work_date"]),
                "shift": normalize_shift_name(raw_entry.get("shift") or "Schicht"),
                "start_time": str(raw_entry.get("start_time") or "")[:5],
                "end_time": str(raw_entry.get("end_time") or "")[:5],
                "notes": str(raw_entry.get("notes") or "")[:500],
            },
            None,
        )
    except (KeyError, TypeError, ValueError):
        return None, "entries require employee_id, work_date and valid values"


def employees_for_entries(entries):
    """Return employee records referenced by entries."""
    employee_ids = sorted(
        {entry_employee_id(entry) for entry in entries if entry_employee_id(entry)}
    )
    if not employee_ids:
        return []
    return Employee.query.filter(Employee.id.in_(employee_ids)).all()


def entry_employee_id(entry):
    """Return the employee id from a model or dict entry."""
    return entry.employee_id if isinstance(entry, ShiftPlanEntry) else entry.get("employee_id")


def conflict_summary(conflicts):
    """Return grouped conflict counts."""
    summary = {"total": len(conflicts), "critical": 0, "warning": 0, "by_type": {}}
    for conflict in conflicts:
        severity = conflict.get("severity")
        if severity in summary:
            summary[severity] += 1
        conflict_type = conflict.get("type") or "unknown"
        summary["by_type"][conflict_type] = summary["by_type"].get(conflict_type, 0) + 1
    return summary


def export_shiftplan_xlsx(plan):
    """Return an XLSX workbook for a shift plan."""
    conflicts_payload = conflicts_for_plan(plan)
    try:
        return export_shiftplan_with_openpyxl(plan, conflicts_payload)
    except ImportError:
        return export_shiftplan_with_stdlib(plan, conflicts_payload)


def export_shiftplan_with_openpyxl(plan, conflicts_payload):
    """Return XLSX bytes using openpyxl when the dependency is installed."""
    from openpyxl import Workbook

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Plan"
    append_rows(
        overview.append,
        plan_metadata_rows(plan)
        + [[""]]
        + [["Datum", "Schicht", "Beginn", "Ende", "Mitarbeiter", "Maschine", "Notiz"]]
        + plan_entry_rows(plan),
    )
    conflicts = workbook.create_sheet("Konflikte")
    append_rows(
        conflicts.append,
        [["Typ", "Schwere", "Datum", "Mitarbeiter", "Maschine", "Meldung"]]
        + conflict_rows(conflicts_payload["conflicts"]),
    )
    summary = workbook.create_sheet("Auswertung")
    append_rows(summary.append, summary_rows(conflicts_payload))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def append_rows(append, rows):
    """Append rows through a sheet append callable."""
    for row in rows:
        append(row)


def export_shiftplan_with_stdlib(plan, conflicts_payload):
    """Return a minimal XLSX workbook using only the standard library."""
    sheets = [
        (
            "Plan",
            plan_metadata_rows(plan)
            + [[""]]
            + [["Datum", "Schicht", "Beginn", "Ende", "Mitarbeiter", "Maschine", "Notiz"]]
            + plan_entry_rows(plan),
        ),
        (
            "Konflikte",
            [["Typ", "Schwere", "Datum", "Mitarbeiter", "Maschine", "Meldung"]]
            + conflict_rows(conflicts_payload["conflicts"]),
        ),
        ("Auswertung", summary_rows(conflicts_payload)),
    ]
    stream = BytesIO()
    with ZipFile(stream, "w", ZIP_DEFLATED) as workbook:
        workbook.writestr("[Content_Types].xml", xlsx_content_types(len(sheets)))
        workbook.writestr("_rels/.rels", xlsx_root_rels())
        workbook.writestr("xl/workbook.xml", xlsx_workbook(sheets))
        workbook.writestr("xl/_rels/workbook.xml.rels", xlsx_workbook_rels(sheets))
        for index, (_, rows) in enumerate(sheets, start=1):
            workbook.writestr(f"xl/worksheets/sheet{index}.xml", xlsx_sheet(rows))
    return stream.getvalue()


def plan_metadata_rows(plan):
    """Return workbook metadata rows for a shift plan."""
    return [
        ["Titel", plan.title],
        ["Abteilung", plan.department],
        ["Startdatum", plan.start_date.isoformat()],
        ["Tage", plan.days],
        ["Rhythmus", plan.rhythm],
        ["Status", plan.status],
    ]


def plan_entry_rows(plan):
    """Return workbook rows for shift plan entries."""
    return [
        [
            entry.work_date.isoformat(),
            entry.shift,
            entry.start_time,
            entry.end_time,
            entry.employee.name if entry.employee else "",
            entry.machine.name if entry.machine else "",
            entry.notes,
        ]
        for entry in sorted(plan.entries, key=lambda item: (item.work_date, item.shift, item.id))
    ]


def conflict_rows(conflicts):
    """Return workbook rows for conflicts."""
    return [
        [
            conflict.get("type", ""),
            conflict.get("severity", ""),
            conflict.get("work_date", ""),
            conflict.get("employee_id", ""),
            conflict.get("machine_id", ""),
            conflict.get("message", ""),
        ]
        for conflict in conflicts
    ]


def summary_rows(conflicts_payload):
    """Return workbook rows for conflict and coverage summary."""
    summary = conflicts_payload["summary"]
    rows = [
        ["Konflikte gesamt", summary["total"]],
        ["Kritisch", summary["critical"]],
        ["Warnungen", summary["warning"]],
        [""],
        ["Typ", "Anzahl"],
    ]
    rows.extend([key, value] for key, value in sorted(summary["by_type"].items()))
    return rows


def xlsx_content_types(sheet_count):
    """Return XLSX content type metadata."""
    sheet_overrides = "".join(
        f'<Override PartName="/xl/worksheets/sheet{index}.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.'
        'spreadsheetml.worksheet+xml"/>'
        for index in range(1, sheet_count + 1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-'
        'package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        f"{sheet_overrides}</Types>"
    )


def xlsx_root_rels():
    """Return XLSX root relationship metadata."""
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/></Relationships>'
    )


def xlsx_workbook(sheets):
    """Return XLSX workbook metadata for sheets."""
    sheet_tags = "".join(
        f'<sheet name="{escape(name)}" sheetId="{index}" r:id="rId{index}"/>'
        for index, (name, _) in enumerate(sheets, start=1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{sheet_tags}</sheets></workbook>"
    )


def xlsx_workbook_rels(sheets):
    """Return XLSX workbook relationship metadata."""
    rels = "".join(
        f'<Relationship Id="rId{index}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        f'Target="worksheets/sheet{index}.xml"/>'
        for index, _ in enumerate(sheets, start=1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f"{rels}</Relationships>"
    )


def xlsx_sheet(rows):
    """Return one XLSX worksheet XML document."""
    row_tags = []
    for row_index, row in enumerate(rows, start=1):
        cells = []
        for column_index, value in enumerate(row, start=1):
            cell_ref = f"{xlsx_column_name(column_index)}{row_index}"
            cells.append(
                f'<c r="{cell_ref}" t="inlineStr"><is><t>{escape(str(value or ""))}</t></is></c>'
            )
        row_tags.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(row_tags)}</sheetData></worksheet>'
    )


def xlsx_column_name(index):
    """Return an Excel column name for a one-based index."""
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def calendar_entries_for_user(user, employee_id=None, start_date=None, days=14, plan_id=None):
    """Return calendar entries for one employee and visible shift plans."""
    try:
        parsed_start_date = parse_date(start_date)
        parsed_days = parse_days(days)
        target_employee_id = int(employee_id) if employee_id not in (None, "") else None
    except (TypeError, ValueError) as exc:
        return None, {"error": str(exc)}, 400

    if not target_employee_id:
        if not user.employee_id:
            return (
                {
                    "employee": None,
                    "start_date": parsed_start_date.isoformat(),
                    "days": parsed_days,
                    "entries": [],
                    "message": "Kein Mitarbeiter mit diesem Benutzer verknuepft.",
                },
                None,
                200,
            )
        target_employee_id = user.employee_id

    if not can_read_calendar_for_employee(user, target_employee_id):
        return None, {"error": "Forbidden"}, 403

    employee = db.session.get(Employee, target_employee_id)
    if not employee:
        return None, {"error": "Mitarbeiter nicht gefunden"}, 404

    end_date = parsed_start_date + timedelta(days=parsed_days)
    query = ShiftPlanEntry.query.filter(
        ShiftPlanEntry.employee_id == target_employee_id,
        ShiftPlanEntry.work_date >= parsed_start_date,
        ShiftPlanEntry.work_date < end_date,
    )
    if plan_id not in (None, ""):
        try:
            query = query.filter(ShiftPlanEntry.plan_id == int(plan_id))
        except (TypeError, ValueError):
            return None, {"error": "plan_id must be a valid integer"}, 400

    entries = query.order_by(
        ShiftPlanEntry.work_date.asc(),
        ShiftPlanEntry.start_time.asc(),
        ShiftPlanEntry.id.asc(),
    ).all()
    payload_entries = [calendar_entry_payload(entry) for entry in entries]
    occupied_dates = {entry.work_date for entry in entries}
    for day_offset in range(parsed_days):
        current_date = parsed_start_date + timedelta(days=day_offset)
        if current_date in occupied_dates:
            continue
        payload_entries.append(free_day_payload(current_date))

    payload_entries.sort(key=lambda item: (item["work_date"], item["start_time"]))
    return (
        {
            "employee": employee.to_dict("basic"),
            "start_date": parsed_start_date.isoformat(),
            "days": parsed_days,
            "entries": payload_entries,
        },
        None,
        200,
    )


def can_read_calendar_for_employee(user, employee_id):
    """Return whether a user may read one employee calendar."""
    if user.is_admin:
        return True
    return user.employee_id == employee_id or has_employee_access(user, "shift")


def calendar_entry_payload(entry):
    """Return one serialized calendar entry with frontend color metadata."""
    shift = normalize_shift_name(entry.shift)
    return {
        "id": entry.id,
        "plan_id": entry.plan_id,
        "work_date": entry.work_date.isoformat(),
        "shift": shift,
        "start_time": entry.start_time,
        "end_time": entry.end_time,
        "machine": entry.machine.to_dict() if entry.machine else None,
        "notes": entry.notes,
        "color": shift_color(shift),
    }


def free_day_payload(work_date):
    """Return a derived free-day calendar entry."""
    return {
        "id": None,
        "plan_id": None,
        "work_date": work_date.isoformat(),
        "shift": "Frei",
        "start_time": "",
        "end_time": "",
        "machine": None,
        "notes": "Frei",
        "color": shift_color("Frei"),
    }


def shift_color(shift):
    """Return the configured calendar color key for a shift name."""
    colors = {
        "Frueh": "green",
        "Spaet": "blue",
        "Nacht": "red",
        "Frei": "violet",
        "Urlaub": "amber",
    }
    return colors.get(normalize_shift_name(shift), "slate")
