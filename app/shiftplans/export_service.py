"""Shift plan XLSX export helpers."""

from io import BytesIO
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

from app.shiftplans.services import conflicts_for_plan


def export_shiftplan_xlsx(plan):
    """Return an XLSX workbook for a shift plan."""
    conflicts_payload = conflicts_for_plan(plan)
    try:
        return export_shiftplan_with_openpyxl(plan, conflicts_payload)
    except ImportError:
        return export_shiftplan_with_stdlib(plan, conflicts_payload)


def export_shiftplan_with_openpyxl(plan, conflicts_payload):
    """Return XLSX bytes using openpyxl when the dependency is installed."""
    from openpyxl import Workbook

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Plan"
    append_rows(
        overview.append,
        plan_metadata_rows(plan)
        + [[""]]
        + [["Datum", "Schicht", "Beginn", "Ende", "Mitarbeiter", "Maschine", "Notiz"]]
        + plan_entry_rows(plan),
    )
    conflicts = workbook.create_sheet("Konflikte")
    append_rows(
        conflicts.append,
        [["Typ", "Schwere", "Datum", "Mitarbeiter", "Maschine", "Meldung"]]
        + conflict_rows(conflicts_payload["conflicts"]),
    )
    summary = workbook.create_sheet("Auswertung")
    append_rows(summary.append, summary_rows(conflicts_payload))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def append_rows(append, rows):
    """Append rows through a sheet append callable."""
    for row in rows:
        append(row)


def export_shiftplan_with_stdlib(plan, conflicts_payload):
    """Return a minimal XLSX workbook using only the standard library."""
    sheets = [
        (
            "Plan",
            plan_metadata_rows(plan)
            + [[""]]
            + [["Datum", "Schicht", "Beginn", "Ende", "Mitarbeiter", "Maschine", "Notiz"]]
            + plan_entry_rows(plan),
        ),
        (
            "Konflikte",
            [["Typ", "Schwere", "Datum", "Mitarbeiter", "Maschine", "Meldung"]]
            + conflict_rows(conflicts_payload["conflicts"]),
        ),
        ("Auswertung", summary_rows(conflicts_payload)),
    ]
    stream = BytesIO()
    with ZipFile(stream, "w", ZIP_DEFLATED) as workbook:
        workbook.writestr("[Content_Types].xml", xlsx_content_types(len(sheets)))
        workbook.writestr("_rels/.rels", xlsx_root_rels())
        workbook.writestr("xl/workbook.xml", xlsx_workbook(sheets))
        workbook.writestr("xl/_rels/workbook.xml.rels", xlsx_workbook_rels(sheets))
        for index, (_, rows) in enumerate(sheets, start=1):
            workbook.writestr(f"xl/worksheets/sheet{index}.xml", xlsx_sheet(rows))
    return stream.getvalue()


def plan_metadata_rows(plan):
    """Return workbook metadata rows for a shift plan."""
    return [
        ["Titel", plan.title],
        ["Abteilung", plan.department],
        ["Startdatum", plan.start_date.isoformat()],
        ["Tage", plan.days],
        ["Rhythmus", plan.rhythm],
        ["Status", plan.status],
    ]


def plan_entry_rows(plan):
    """Return workbook rows for shift plan entries."""
    return [
        [
            entry.work_date.isoformat(),
            entry.shift,
            entry.start_time,
            entry.end_time,
            entry.employee.name if entry.employee else "",
            entry.machine.name if entry.machine else "",
            entry.notes,
        ]
        for entry in sorted(plan.entries, key=lambda item: (item.work_date, item.shift, item.id))
    ]


def conflict_rows(conflicts):
    """Return workbook rows for conflicts."""
    return [
        [
            conflict.get("type", ""),
            conflict.get("severity", ""),
            conflict.get("work_date", ""),
            conflict.get("employee_id", ""),
            conflict.get("machine_id", ""),
            conflict.get("message", ""),
        ]
        for conflict in conflicts
    ]


def summary_rows(conflicts_payload):
    """Return workbook rows for conflict and coverage summary."""
    summary = conflicts_payload["summary"]
    rows = [
        ["Konflikte gesamt", summary["total"]],
        ["Kritisch", summary["critical"]],
        ["Warnungen", summary["warning"]],
        [""],
        ["Typ", "Anzahl"],
    ]
    rows.extend([key, value] for key, value in sorted(summary["by_type"].items()))
    return rows


def xlsx_content_types(sheet_count):
    """Return XLSX content type metadata."""
    sheet_overrides = "".join(
        f'<Override PartName="/xl/worksheets/sheet{index}.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.'
        'spreadsheetml.worksheet+xml"/>'
        for index in range(1, sheet_count + 1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-'
        'package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        f"{sheet_overrides}</Types>"
    )


def xlsx_root_rels():
    """Return XLSX root relationship metadata."""
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/></Relationships>'
    )


def xlsx_workbook(sheets):
    """Return XLSX workbook metadata for sheets."""
    sheet_tags = "".join(
        f'<sheet name="{escape(name)}" sheetId="{index}" r:id="rId{index}"/>'
        for index, (name, _) in enumerate(sheets, start=1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{sheet_tags}</sheets></workbook>"
    )


def xlsx_workbook_rels(sheets):
    """Return XLSX workbook relationship metadata."""
    rels = "".join(
        f'<Relationship Id="rId{index}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        f'Target="worksheets/sheet{index}.xml"/>'
        for index, _ in enumerate(sheets, start=1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f"{rels}</Relationships>"
    )


def xlsx_sheet(rows):
    """Return one XLSX worksheet XML document."""
    row_tags = []
    for row_index, row in enumerate(rows, start=1):
        cells = []
        for column_index, value in enumerate(row, start=1):
            cell_ref = f"{xlsx_column_name(column_index)}{row_index}"
            cells.append(
                f'<c r="{cell_ref}" t="inlineStr"><is><t>{escape(str(value or ""))}</t></is></c>'
            )
        row_tags.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(row_tags)}</sheetData></worksheet>'
    )


def xlsx_column_name(index):
    """Return an Excel column name for a one-based index."""
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name
